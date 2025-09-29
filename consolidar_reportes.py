import pandas as pd
import os
from openpyxl.utils import get_column_letter
import numpy as np

# --- Configuración ---
carpeta_reportes = 'reportes_mensuales'
archivo_salida = 'resumen_consolidado.xlsx'
nombre_de_la_hoja = 'REPORTE FACTURACIÓN' 

# --- AJUSTES DE LECTURA DE LA TABLA ---
filas_a_saltar_al_inicio = 2 

# --- INICIO DE LA LÓGICA DEL SCRIPT ---
lista_de_datos = []
nombres_columnas_finales = [
    'N°', 'Ciudad', 'Agencia', 'Ubicación', 'Departamento', 'Modelo', 'Serial', 'IP',
    'Inicial Imp. B/N', 'Final Imp. B/N', 'Total Imp. B/N',
    'Inicial Copias B/N', 'Final Copias B/N', 'Total Copias B/N',
    'Inicial Imp. Color', 'Final Imp. Color', 'Total Imp. Color',
    'Inicial Copias Color', 'Final Copias Color', 'Total Copias Color'
]

print(f"Buscando archivos de Excel en la carpeta: '{carpeta_reportes}'...")

for archivo in os.listdir(carpeta_reportes):
    if archivo.endswith(('.xlsx', '.xls')) and not archivo.startswith('~$'):
        ruta_completa = os.path.join(carpeta_reportes, archivo)
        print(f"Leyendo la tabla de datos del archivo: {archivo}...")
        try:
            df_mensual = pd.read_excel(
                ruta_completa,
                sheet_name=nombre_de_la_hoja,
                skiprows=filas_a_saltar_al_inicio
            )
            # Eliminar filas donde la columna 'Serial' está vacía para asegurar datos válidos.
            df_mensual.dropna(subset=['Serial'], inplace=True)
            
            # Lógica adaptativa para manejar el formato nuevo (con columna extra)
            if len(df_mensual.columns) > 20:
                df_mensual.drop(df_mensual.columns[7], axis=1, inplace=True) # Elimina columna en blanco
            
            # Asegurar que solo se usen las 20 columnas de datos
            df_mensual = df_mensual.iloc[:, :20]
            df_mensual.columns = nombres_columnas_finales
            
            nombre_periodo = os.path.splitext(archivo)[0] 
            df_mensual['Periodo'] = nombre_periodo
            lista_de_datos.append(df_mensual)
            
        except Exception as e:
            print(f"  -> ADVERTENCIA: No se pudo leer el archivo '{archivo}'. Error: {e}")

if not lista_de_datos:
    print("Error: No se procesó ningún archivo con éxito.")
else:
    print("Consolidando todos los datos...")
    df_consolidado = pd.concat(lista_de_datos, ignore_index=True)
    
    print("Ordenando los datos cronológicamente...")
    meses_es = {
        'Enero': 'January', 'Febrero': 'February', 'Marzo': 'March', 'Abril': 'April', 'Mayo': 'May', 'Junio': 'June',
        'Julio': 'July', 'Agosto': 'August', 'Septiembre': 'September', 'Octubre': 'October', 'Noviembre': 'November', 'Diciembre': 'December'
    }
    # Extraer mes y año de forma segura
    periodo_partes = df_consolidado['Periodo'].str.split()
    periodo_traducido = periodo_partes.str[0].map(meses_es)
    año_str = periodo_partes.str[1]
    
    periodo_en_ingles = periodo_traducido + ' ' + año_str
    df_consolidado['FechaOrden'] = pd.to_datetime(periodo_en_ingles, errors='coerce')
    df_consolidado = df_consolidado.sort_values('FechaOrden').drop(columns=['FechaOrden'])

    print("Limpiando datos y recalculando totales...")
    columnas_contadores = [
        'Inicial Imp. B/N', 'Final Imp. B/N', 'Inicial Copias B/N', 'Final Copias B/N',
        'Inicial Imp. Color', 'Final Imp. Color', 'Inicial Copias Color', 'Final Copias Color'
    ]
    for col in columnas_contadores:
        # Convierte a numérico, los errores se vuelven NaN, y luego se rellenan con 0
        df_consolidado[col] = pd.to_numeric(df_consolidado[col], errors='coerce').fillna(0)

    # Recalcula los totales para asegurar consistencia
    df_consolidado['Total Imp. B/N'] = np.where(df_consolidado['Final Imp. B/N'] >= df_consolidado['Inicial Imp. B/N'], df_consolidado['Final Imp. B/N'] - df_consolidado['Inicial Imp. B/N'], df_consolidado['Final Imp. B/N'])
    df_consolidado['Total Copias B/N'] = np.where(df_consolidado['Final Copias B/N'] >= df_consolidado['Inicial Copias B/N'], df_consolidado['Final Copias B/N'] - df_consolidado['Inicial Copias B/N'], df_consolidado['Final Copias B/N'])
    df_consolidado['Total Imp. Color'] = np.where(df_consolidado['Final Imp. Color'] >= df_consolidado['Inicial Imp. Color'], df_consolidado['Final Imp. Color'] - df_consolidado['Inicial Imp. Color'], df_consolidado['Final Imp. Color'])
    df_consolidado['Total Copias Color'] = np.where(df_consolidado['Final Copias Color'] >= df_consolidado['Inicial Copias Color'], df_consolidado['Final Copias Color'] - df_consolidado['Inicial Copias Color'], df_consolidado['Final Copias Color'])
    
    # --- AJUSTE: Convertir columnas numéricas a enteros para evitar decimales ---
    print("Ajustando formato de columnas numéricas a entero...")
    columnas_a_convertir_entero = [
        'Inicial Imp. B/N', 'Final Imp. B/N', 'Total Imp. B/N',
        'Inicial Copias B/N', 'Final Copias B/N', 'Total Copias B/N',
        'Inicial Imp. Color', 'Final Imp. Color', 'Total Imp. Color',
        'Inicial Copias Color', 'Final Copias Color', 'Total Copias Color'
    ]
    for col in columnas_a_convertir_entero:
        df_consolidado[col] = df_consolidado[col].astype(np.int64)

    # --- AJUSTE: Convertir la columna de IP a texto ---
    print("Cambiando formato de columna IP a texto...")
    df_consolidado['IP'] = df_consolidado['IP'].astype(str)

    print("Calculando totales anuales...")
    resumen_anual = {'Concepto': [], 'Total Anual': []}
    columnas_totales = ['Total Imp. B/N', 'Total Copias B/N', 'Total Imp. Color', 'Total Copias Color']
    for col in columnas_totales:
        resumen_anual['Concepto'].append(col.replace('Total', 'Total de'))
        resumen_anual['Total Anual'].append(df_consolidado[col].sum())
    df_resumen = pd.DataFrame(resumen_anual)

    print("Guardando el archivo final con dos hojas...")
    with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
        df_consolidado.to_excel(writer, index=False, sheet_name='Resumen Detallado')
        df_resumen.to_excel(writer, index=False, sheet_name='Totales Anuales')
        
        # Auto-ajustar ancho de columnas en la primera hoja
        worksheet1 = writer.sheets['Resumen Detallado']
        for column_cells in worksheet1.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = max_length + 2
            worksheet1.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

        # Auto-ajustar ancho de columnas en la segunda hoja
        worksheet2 = writer.sheets['Totales Anuales']
        for column_cells in worksheet2.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = max_length + 2
            worksheet2.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
    
    print("-" * 30)
    print(f"✅ ¡Proceso completado! El archivo está guardado en: '{archivo_salida}'")